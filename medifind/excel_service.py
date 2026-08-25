import io
import csv
import logging
from datetime import datetime, date
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from django.utils import timezone
from django.db import transaction

from .models import Medicine, Inventory, Pharmacy

logger = logging.getLogger(__name__)


class ExcelInventoryService:
    """
    Handles Excel/CSV inventory template generation, bulk file importing,
    data validation, and active inventory export.
    """

    TEMPLATE_COLUMNS = [
        ("Medicine Name", "Dolo 650", "Full generic or trade medicine name (Required)"),
        ("Brand", "Micro Labs", "Pharmaceutical manufacturer / brand"),
        ("Category", "Pain Relief", "e.g. Pain Relief, Fever & Cold, Antibiotic, Diabetes Care"),
        ("Dosage", "650mg", "Dosage strength e.g. 500mg, 10mg, 5ml"),
        ("Price (INR)", "30.50", "Retail price per unit/pack in INR (Required)"),
        ("Quantity (Stock)", "100", "Current stock quantity in store (Required)"),
        ("Batch Number", "DOLO-2026-X1", "Manufacturer batch identifier"),
        ("Expiry Date", "2027-12-31", "Format: YYYY-MM-DD or DD/MM/YYYY (Required)"),
        ("Package Size", "Strip of 15", "e.g. Strip of 15, Strip of 10, Bottle of 100 ml, Unit"),
        ("SKU Code", "DOLO650-S15", "Unique store SKU barcode/code"),
        ("Description", "Effective paracetamol tablet for pain and fever relief", "Clinical summary"),
        ("Uses", "Fever, Headache, Body Pain", "Conditions treated"),
        ("Side Effects", "Rarely causes nausea or dizziness", "Noted side effects"),
        ("Prescription Required", "No", "Yes or No"),
    ]

    SAMPLE_ROWS = [
        [
            "Dolo 650", "Micro Labs", "Pain Relief", "650mg",
            30.50, 150, "DOLO-2026-A1", "2027-08-31", "Strip of 15",
            "DOLO650-S15", "Fast-acting antipyretic and analgesic", "Fever, Body ache", "None under standard dosage", "No"
        ],
        [
            "Azithromycin 500mg", "Zithrocare", "Antibiotic", "500mg",
            120.00, 45, "AZI-2026-M4", "2027-04-15", "Strip of 10",
            "AZI500-S10", "Broad-spectrum macrolide antibiotic", "Bacterial respiratory infections", "Mild stomach upset", "Yes"
        ],
        [
            "Paracetamol 500mg", "Calpol", "Fever & Cold", "500mg",
            18.00, 200, "CAL-2026-09", "2027-11-30", "Strip of 15",
            "CAL500-S15", "Effective pain relief and fever reducer", "Cold, Headache, Fever", "Safe under prescribed limit", "No"
        ],
        [
            "Metformin 500mg", "Glycomet", "Diabetes Care", "500mg",
            42.00, 80, "GLY-2026-88", "2028-01-20", "Strip of 10",
            "GLY500-S10", "First-line medication for type 2 diabetes", "Blood sugar management", "Mild gastrointestinal effect", "Yes"
        ],
        [
            "Pantoprazole 40mg", "Pan 40", "Digestive Health", "40mg",
            95.00, 90, "PAN-2026-33", "2027-10-31", "Strip of 15",
            "PAN40-S15", "Proton pump inhibitor reducing stomach acid", "Acidity, GERD, Heartburn", "Headache in rare cases", "No"
        ],
        [
            "Cetirizine 10mg", "Cetzine", "Allergy", "10mg",
            22.00, 120, "CET-2026-72", "2027-09-15", "Strip of 10",
            "CET10-S10", "Second-generation antihistamine", "Allergic rhinitis, Sneezing, Hives", "Mild drowsiness", "No"
        ],
    ]

    @classmethod
    def generate_excel_template(cls):
        """
        Generates a beautifully styled .xlsx inventory template workbook.
        Returns bytes buffer.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventory_Upload_Template"
        ws.views.sheetView[0].showGridLines = True

        # Styles
        header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid") # Teal-600
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        sub_fill = PatternFill(start_color="F0FDFA", end_color="F0FDFA", fill_type="solid") # Teal-50
        sub_font = Font(name="Segoe UI", size=9, italic=True, color="0F766E")
        data_font = Font(name="Segoe UI", size=10, color="1E293B")
        zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )

        # Row 1: Headers
        for col_idx, (col_name, _, _) in enumerate(cls.TEMPLATE_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[1].height = 28

        # Row 2: Helper descriptions
        for col_idx, (_, _, desc) in enumerate(cls.TEMPLATE_COLUMNS, 1):
            cell = ws.cell(row=2, column=col_idx, value=desc)
            cell.fill = sub_fill
            cell.font = sub_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[2].height = 24

        # Sample data rows
        for row_idx, sample_row in enumerate(cls.SAMPLE_ROWS, 3):
            is_zebra = (row_idx % 2 == 0)
            for col_idx, val in enumerate(sample_row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = data_font
                cell.border = thin_border
                if is_zebra:
                    cell.fill = zebra_fill
                if isinstance(val, (int, float, Decimal)):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row_idx].height = 20

        # Auto-adjust column widths
        for col_idx, (col_name, _, _) in enumerate(cls.TEMPLATE_COLUMNS, 1):
            col_letter = get_column_letter(col_idx)
            max_len = max(len(str(col_name)), 14)
            ws.column_dimensions[col_letter].width = max_len + 5

        # Instructions sheet
        info_ws = wb.create_sheet(title="Instructions_&_Guidelines")
        info_ws.views.sheetView[0].showGridLines = True
        info_title_font = Font(name="Segoe UI", size=13, bold=True, color="0F766E")
        info_body_font = Font(name="Segoe UI", size=10, color="334155")
        
        guidelines = [
            ("MediAI Pharmacy Inventory Upload Guidelines", info_title_font),
            ("", info_body_font),
            ("1. How it Works:", Font(name="Segoe UI", size=11, bold=True, color="0F172A")),
            ("   • Fill out your inventory in the 'Inventory_Upload_Template' sheet.", info_body_font),
            ("   • You can keep or delete the sample rows before uploading.", info_body_font),
            ("   • When you upload, MediAI updates matching medicines and adds new stock instantly.", info_body_font),
            ("", info_body_font),
            ("2. Mandatory Fields:", Font(name="Segoe UI", size=11, bold=True, color="0F172A")),
            ("   • Medicine Name (e.g. Dolo 650)", info_body_font),
            ("   • Price (INR) (Retail selling price per pack)", info_body_font),
            ("   • Quantity (Stock) (Available units in your pharmacy)", info_body_font),
            ("   • Expiry Date (YYYY-MM-DD or DD/MM/YYYY)", info_body_font),
            ("", info_body_font),
            ("3. Supported Categories:", Font(name="Segoe UI", size=11, bold=True, color="0F172A")),
            ("   • Pain Relief, Fever & Cold, Allergy, Digestive Health, Vitamins & Supplements", info_body_font),
            ("   • Diabetes Care, Blood Pressure, Skin Care, First Aid, Respiratory, Antibiotic, Heart", info_body_font),
            ("", info_body_font),
            ("4. Real-Time API Integration Note:", Font(name="Segoe UI", size=11, bold=True, color="0F172A")),
            ("   • If your pharmacy uses a POS/ERP system with an API, you can enable Real-Time API Sync", info_body_font),
            ("     in your Pharmacy Dashboard so customer searches query your system directly!", info_body_font),
        ]

        for r_idx, (text, f_style) in enumerate(guidelines, 1):
            c = info_ws.cell(row=r_idx, column=1, value=text)
            c.font = f_style
        info_ws.column_dimensions["A"].width = 95

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @classmethod
    def generate_csv_template(cls):
        """
        Generates a standard CSV inventory template.
        """
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([col[0] for col in cls.TEMPLATE_COLUMNS])
        for row in cls.SAMPLE_ROWS:
            writer.writerow(row)
        return output.getvalue()

    @classmethod
    def export_pharmacy_inventory(cls, pharmacy):
        """
        Exports the current active inventory of a pharmacy to an Excel workbook.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Inventory_{pharmacy.name[:20]}"
        ws.views.sheetView[0].showGridLines = True

        header_fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="Segoe UI", size=10, color="1E293B")
        zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )

        headers = [col[0] for col in cls.TEMPLATE_COLUMNS]
        for col_idx, col_name in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        ws.row_dimensions[1].height = 26

        inventory_qs = Inventory.objects.filter(pharmacy=pharmacy).select_related("medicine").order_by("medicine__name")
        for row_idx, item in enumerate(inventory_qs, 2):
            is_zebra = (row_idx % 2 == 0)
            med = item.medicine
            row_data = [
                med.name,
                med.brand,
                med.category,
                med.dosage,
                float(item.price),
                item.quantity,
                item.batch_number or "",
                item.expiry_date.strftime("%Y-%m-%d") if item.expiry_date else "",
                item.package_size or "Strip of 15",
                item.sku_code or "",
                med.description,
                med.uses,
                med.side_effects,
                "Yes" if med.prescription_required else "No"
            ]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = data_font
                cell.border = thin_border
                if is_zebra:
                    cell.fill = zebra_fill
                if isinstance(val, (int, float, Decimal)):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row_idx].height = 20

        for col_idx, col_name in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(len(col_name), 14) + 4

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    MAX_UPLOAD_SIZE_BYTES = 5 * 1024 * 1024  # 5 MB limit
    MAX_ROWS_LIMIT = 5000  # Prevent memory exhaustion / decompression bombs
    ALLOWED_EXTENSIONS = ('.xlsx', '.xls', '.csv')

    @classmethod
    def import_inventory_file(cls, pharmacy, file_obj, filename=""):
        """
        Parses an uploaded Excel (.xlsx, .xls) or CSV file and batch-syncs medicines and inventory.
        Enforces file size limit, allowed extensions, and maximum row count.
        """
        filename_lower = (filename or getattr(file_obj, "name", "")).lower()

        # 1. Extension validation
        if not any(filename_lower.endswith(ext) for ext in cls.ALLOWED_EXTENSIONS):
            return {"success": False, "message": "Invalid file format. Only .xlsx, .xls, and .csv files are supported."}

        # 2. File size validation
        size = getattr(file_obj, "size", None)
        if size and size > cls.MAX_UPLOAD_SIZE_BYTES:
            return {"success": False, "message": f"File exceeds maximum allowed upload size of 5 MB."}

        if filename_lower.endswith(".csv"):
            return cls._import_from_csv(pharmacy, file_obj)
        else:
            return cls._import_from_excel(pharmacy, file_obj)

    @classmethod
    def _import_from_excel(cls, pharmacy, file_obj):
        try:
            wb = openpyxl.load_workbook(file_obj, data_only=True, read_only=False)
            # Pick first non-instruction sheet
            ws = wb.active
            for sheet in wb.worksheets:
                if "instruction" not in sheet.title.lower():
                    ws = sheet
                    break

            rows = []
            for idx, row in enumerate(ws.iter_rows(values_only=True)):
                if idx >= cls.MAX_ROWS_LIMIT:
                    logger.warning(f"Excel import exceeded maximum {cls.MAX_ROWS_LIMIT} rows limit.")
                    break
                rows.append(row)

            if not rows:
                return {"success": False, "message": "The uploaded spreadsheet is empty."}

            return cls._process_parsed_rows(pharmacy, rows)
        except Exception as e:
            logger.error(f"Excel import parsing error: {e}", exc_info=True)
            return {"success": False, "message": f"Failed to parse Excel file: {str(e)}"}


    @classmethod
    def _import_from_csv(cls, pharmacy, file_obj):
        try:
            content = file_obj.read()
            if isinstance(content, bytes):
                try:
                    text = content.decode("utf-8-sig")
                except UnicodeDecodeError:
                    text = content.decode("latin-1")
            else:
                text = str(content)

            reader = csv.reader(io.StringIO(text))
            rows = list(reader)
            if not rows:
                return {"success": False, "message": "The uploaded CSV file is empty."}

            return cls._process_parsed_rows(pharmacy, rows)
        except Exception as e:
            logger.error(f"CSV import parsing error: {e}", exc_info=True)
            return {"success": False, "message": f"Failed to parse CSV file: {str(e)}"}

    @classmethod
    def _process_parsed_rows(cls, pharmacy, raw_rows):
        """
        Normalizes header columns, validates medicine & inventory data,
        and saves records atomically with comprehensive validation logging.
        """
        if len(raw_rows) < 2:
            return {"success": False, "message": "The file contains headers but no data rows."}

        # Locate header row (find row containing 'medicine' or 'name')
        header_idx = 0
        header_map = {}
        for idx, row in enumerate(raw_rows[:5]):
            if not row:
                continue
            normalized = [str(c or "").strip().lower() for c in row]
            if any("medicine" in c or "name" in c or "drug" in c for c in normalized):
                header_idx = idx
                for c_idx, col_name in enumerate(normalized):
                    col_clean = col_name.replace(" ", "").replace("_", "").replace("-", "").replace("(inr)", "").replace("(stock)", "")
                    header_map[col_clean] = c_idx
                break

        if not header_map:
            # Fallback to row 0
            for c_idx, c_val in enumerate(raw_rows[0]):
                col_clean = str(c_val or "").strip().lower().replace(" ", "").replace("_", "")
                header_map[col_clean] = c_idx

        def get_col(row, *aliases):
            for alias in aliases:
                norm_alias = alias.replace(" ", "").replace("_", "").replace("-", "")
                if norm_alias in header_map:
                    col_idx = header_map[norm_alias]
                    if col_idx < len(row):
                        val = row[col_idx]
                        if val is not None and str(val).strip() != "":
                            return val
            return None

        created_count = 0
        updated_count = 0
        errors = []
        data_rows = raw_rows[header_idx + 1:]

        with transaction.atomic():
            for row_num, row in enumerate(data_rows, start=header_idx + 2):
                if not row or all(c is None or str(c).strip() == "" for c in row):
                    continue

                # Skip secondary explanation rows (e.g. "format: yyyy-mm-dd")
                first_val = str(row[0] or "").strip().lower()
                if "format:" in first_val or "required" in first_val or "e.g." in first_val:
                    continue

                med_name = get_col(row, "medicinename", "medicine", "name", "drugname", "productname")
                if not med_name:
                    errors.append(f"Row {row_num}: Missing Medicine Name.")
                    continue
                med_name = str(med_name).strip()

                brand = str(get_col(row, "brand", "manufacturer", "mfg", "company") or "Generic").strip()
                category = str(get_col(row, "category", "type", "class") or "General Health").strip()
                dosage = str(get_col(row, "dosage", "strength", "dose") or "Standard").strip()
                description = str(get_col(row, "description", "desc") or f"{med_name} {dosage}").strip()
                uses = str(get_col(row, "uses", "indications") or "General Treatment").strip()
                side_effects = str(get_col(row, "sideeffects", "sideeffect") or "Consult physician").strip()
                
                raw_rx = str(get_col(row, "prescriptionrequired", "prescription", "rx") or "no").strip().lower()
                prescription_required = raw_rx in ("yes", "true", "1", "y", "req", "required")

                # Price
                raw_price = get_col(row, "price", "priceinr", "mrp", "unitprice", "rate", "cost")
                if raw_price is None:
                    errors.append(f"Row {row_num} ({med_name}): Missing price.")
                    continue
                try:
                    price = Decimal(str(raw_price).replace("₹", "").replace(",", "").strip())
                    if price < 0:
                        raise ValueError("Price cannot be negative")
                except Exception:
                    errors.append(f"Row {row_num} ({med_name}): Invalid price value '{raw_price}'.")
                    continue

                # Quantity
                raw_qty = get_col(row, "quantity", "quantitystock", "stock", "qty", "units")
                if raw_qty is None:
                    raw_qty = 0
                try:
                    quantity = int(float(str(raw_qty).replace(",", "").strip()))
                    if quantity < 0:
                        quantity = 0
                except Exception:
                    errors.append(f"Row {row_num} ({med_name}): Invalid quantity value '{raw_qty}'.")
                    continue

                # Batch Number
                batch_number = str(get_col(row, "batchnumber", "batch", "batchno", "lot") or f"EXCEL-{timezone.now().strftime('%m%y')}").strip()

                # Package Size
                package_size = str(get_col(row, "packagesize", "packsize", "pack", "packaging") or "Strip of 15").strip()

                # SKU Code
                sku_code = get_col(row, "skucode", "sku", "itemcode", "barcode")
                if sku_code:
                    sku_code = str(sku_code).strip()
                else:
                    clean_name = ''.join(c for c in med_name if c.isalnum())[:8].upper()
                    sku_code = f"{clean_name}-{package_size[:4].replace(' ', '').upper()}"

                # Expiry Date
                raw_exp = get_col(row, "expirydate", "expiry", "expdate", "exp")
                exp_date = cls._parse_date(raw_exp)
                if not exp_date:
                    exp_date = timezone.now().date() + timezone.timedelta(days=365)

                # 1. Upsert Medicine
                medicine, _ = Medicine.objects.get_or_create(
                    name=med_name,
                    defaults={
                        "brand": brand,
                        "category": category,
                        "dosage": dosage,
                        "description": description,
                        "uses": uses,
                        "side_effects": side_effects,
                        "prescription_required": prescription_required,
                    }
                )

                # 2. Upsert Inventory
                existing_inv = Inventory.objects.filter(
                    pharmacy=pharmacy,
                    medicine=medicine,
                    package_size=package_size
                ).first()

                if existing_inv:
                    existing_inv.quantity = quantity
                    existing_inv.price = price
                    existing_inv.batch_number = batch_number
                    existing_inv.expiry_date = exp_date
                    existing_inv.sku_code = sku_code
                    existing_inv.save()
                    updated_count += 1
                else:
                    Inventory.objects.create(
                        pharmacy=pharmacy,
                        medicine=medicine,
                        package_size=package_size,
                        quantity=quantity,
                        price=price,
                        batch_number=batch_number,
                        expiry_date=exp_date,
                        sku_code=sku_code
                    )
                    created_count += 1

        total_processed = created_count + updated_count
        return {
            "success": True,
            "total_processed": total_processed,
            "created_count": created_count,
            "updated_count": updated_count,
            "error_count": len(errors),
            "errors": errors[:15],
            "message": f"Successfully processed {total_processed} items ({created_count} added, {updated_count} updated)."
        }

    @staticmethod
    def _parse_date(val):
        if not val:
            return None
        if isinstance(val, date) and not isinstance(val, datetime):
            return val
        if isinstance(val, datetime):
            return val.date()

        val_str = str(val).strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%Y/%m/%d", "%d.%m.%Y"):
            try:
                return datetime.strptime(val_str, fmt).date()
            except ValueError:
                pass
        return None
